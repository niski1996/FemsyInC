import math

import numpy as np

import programy.mes_obliczenia_1_2.abstr as absc
import programy.mes_obliczenia_1_2.funkcje_pomocnicze as fu_pom
from biblioteki.matr_pom import macierze_num
import programy.mes_obliczenia_1_2.testy.dane_test_main as tst
import programy.mes_obliczenia_1_2.dane_proj as prj
from programy.mes_obliczenia_1_2 import selector

from programy.Visualizer import visualizer
from programy.mes_obliczenia_1_2.funkcje_pomocnicze import swich_poly
from szybkie_do_zmiany import gotowe_k_loc
from openpyxl import Workbook


class MesMain(absc.DeklaracjaZmiennych):
    """Główna klasa programu"""

    def __init__(self, dane, autorun=True, show=True):
        super().__init__()
        self.input_val(dane)
        if autorun:
            self.obl()
        if show:
            self.show()

    def obl(self):
        """Rozpoczyna obliczenia, wynikiem jest macierz sztywności K
        korzysta z tablic węzłów i tablicy elementów """
        self.node_glob = self.num_elem_to_coords()
        self.cos_oldtonew = self.transform_mat()
        self.elem_loc = self.elem_trans()
        self.geom_property()
        # self.elem_loc=self.node_glob#wywalam transformacje

        self.n_loc_element = self.n_loc_element_create()
        self.b = self.n_deriv()
        self.k_e_loc = self.k_e_loc_create()  # dodac w abstrakcie

        if self.el_type == 'surface':  # todo do wyjabania stad
            self.surfacjany()
        self.k_e_glob = self.k_e_glob_create()

        # self.k_e_glob=self.k_e_loc#wywalam transformacje

        self.agregacja()
        self.k_mod_create()

        self.displacement()
        self.reaction()
        self.internal_forces()
        self.inspect_input()

    def surfacjany(self):  # todo do wyjebania
        li = []
        for num, nodes in enumerate(self.elem_loc):
            li.append(gotowe_k_loc(nodes, self.materials[self.elements[num][-2]][0],
                                   self.materials[self.elements[num][-2]][3],
                                   self.materials[self.elements[num][-2]][2]))
        for num_k, k_loc in enumerate(self.k_e_loc):
            e = self.materials[self.elements[num_k][-2]][0]
            a = self.A[num_k]
            h = self.materials[self.elements[num][-2]][3]
            alfa = 0.3
            ks = alfa * e * h * a * np.array([[1, -0.5, -0.5],
                                              [-0.5, 1, -0.5],
                                              [-0.5, -0.5, 1]])

            def do_tarcz(mat):
                li = list(range(2, len(mat) + 2, 2))
                li.pop()
                li.reverse()
                mat = np.c_[mat, np.zeros(len(mat))]
                mat = np.r_[mat, np.zeros((1, len(mat) + 1))]
                mat = np.c_[mat, np.zeros(len(mat))]
                mat = np.r_[mat, np.zeros((1, len(mat) + 1))]
                mat = np.c_[mat, np.zeros(len(mat))]
                mat = np.r_[mat, np.zeros((1, len(mat) + 1))]
                mat = np.c_[mat, np.zeros(len(mat))]
                mat = np.r_[mat, np.zeros((1, len(mat) + 1))]
                for el in li:
                    for nom in range(4):
                        mat = np.insert(mat, el, 0, axis=1)
                        mat = np.insert(mat, el, 0, axis=0)
                return mat

            def do_plyt(mat):
                li = list(range(3, len(mat) + 2, 3))
                li.pop()
                li.reverse()
                mat = np.c_[mat, np.zeros(len(mat))]
                mat = np.r_[mat, np.zeros((1, len(mat) + 1))]
                for el in li:
                    for nom in range(3):
                        mat = np.insert(mat, el, 0, axis=1)
                        mat = np.insert(mat, el, 0, axis=0)
                tmp = np.zeros([18, 18])
                tmp[2:, 2:] = mat
                return tmp

            tmp_pl = do_plyt(li[num_k])
            tmp_sh = do_tarcz(k_loc)
            tmp = tmp_pl + tmp_sh
            tmp[5][5] = ks[0][0]
            tmp[11][5] = ks[1][0]
            tmp[17][5] = ks[2][0]

            tmp[5][11] = ks[0][1]
            tmp[11][11] = ks[1][1]
            tmp[17][11] = ks[2][1]

            tmp[5][17] = ks[0][2]
            tmp[11][17] = ks[1][2]
            tmp[17][17] = ks[2][2]
            self.k_e_loc[num_k] = tmp
        self.dof = 6

    def num_elem_to_coords(self):
        """
        change numer of elements to responding coords
        :return:list
        tablica współrzednych wezłów w ukłądzie globalnym
        lista trzypoziomowa [[(xn,yn,zn),(xm,ym,zm)...],
                            [(),()...],...]
        """
        mac_wsp = []
        for el in self.elements:
            mac_wsp.append([self.nodes[int(el[ind])] for ind in range(len(el) - 2)])
        return mac_wsp

    def transform_mat(self):
        """
        create transformation matrix
        :return: np.array shape n,3,3
        """
        li = []
        for elem in self.node_glob:
            li.append(selector.cos_mat_set(*elem))
        return li

    def elem_trans(self):
        """transpose elements from global to local"""
        for num, elem in enumerate(self.node_glob):
            tmp = []
            for nod in elem:
                c = self.cos_oldtonew[num].T
                col = np.matmul(self.cos_oldtonew[num], np.array(nod).reshape(3, 1))
                tmp.append(list(col.reshape(3)))
            self.elem_loc.append(tmp)
        return self.elem_loc

    def geom_property(self):
        """start metods to count geometric properties"""
        self.V = self.count_Volume()
        self.A = self.count_Area()
        self.D = self.count_D_PSN()
        self.L = self.count_Length()
        self.L = self.count_Length()
        self.EA = self.count_EA()

    # def count_D(self):
    #     li=[]
    #     for num, el in enumerate(self.materials):
    #         v= el[2]
    #         e=el[0]
    #         mnożnik = e/((1+v)*(1-2*v))
    #         vp = 1-v
    #         tp=(1-2*v)/2
    #         mat = np.array([[vp,v,v,0,0,0],
    #                          [v,vp,v,0,0,0],
    #                          [v,v,vp,0,0,0],
    #                          [0,0,0,tp,0,0],
    #                          [0,0,0,0,tp,0],
    #                          [0,0,0,0,0,tp]])
    #         li.append(mnożnik*mat)
    #     return li

    def count_D_PSN(self):
        li = []
        for num, el in enumerate(self.materials):
            v = el[2]
            e = el[0]
            mnoznik = e / (1 - v ** 2)
            vp = 1
            tp = (1 - v) / 2
            mat = np.array([[vp, v, 0],
                            [v, vp, 0],
                            [0, 0, tp]])
            li.append(mnoznik * mat)
        return li

    def count_Area(self):
        li = []
        for num, el in enumerate(self.elements):
            li.append(selector.set_area(el[-1], material=self.materials[el[-2]], elem_loc=self.elem_loc[num]))
        return li

    def count_EA(self):
        li = []
        for num, el in enumerate(self.elements):
            li.append(self.A[num] * self.materials[el[-2]][0])
        return li

    def sztywnosc(self, elem_type, elem_num):
        return selector.stif(elem_type, l=self.L[elem_num], ea=self.EA[elem_num], d=self.D[self.elements[elem_num][-2]],
                             v=self.V[elem_num])

    def count_Length(self):
        li = []
        for num, el in enumerate(self.elem_loc):
            li.append(selector.set_length(self.elements[num][-1], el))
        return li

    def count_Volume(self):
        for num, elem in enumerate(self.elements):
            self.V.append(selector.set_volume(elem[-1], self.elem_loc[num], self.materials[elem[-2]]))
        return self.V

    def n_loc_element_create(self):
        """tworzy macierze funkcji kształtu w układzie lokalnym

        """
        elem_li = []  # lista funkcji kształtu elementów
        for num, elem in enumerate(self.elem_loc):
            inside_li = []  # wewnetrzna lista wielomoianów w elemencie
            val = list(np.zeros(len(elem)))  # wektro wartosci w punktach
            val[0] = 1  # zgodnie z teorią we wszystkich punktach z wyjatkiem jednego przyjmuje 0
            for nodes in elem:
                inside_li.append(selector.set_shape_fuu(self.elements[num][-1], elem, val))
                val.insert(0, val.pop())
            elem_li.append(np.array(inside_li).reshape(1, len(inside_li)))

        return elem_li

    def n_deriv(self):
        li = []
        for num, elem in enumerate(self.n_loc_element):
            li.append(selector.set_deriv_matr(self.elements[num][-1], elem))
        return li

    def k_e_loc_create(self):
        """
        może zawierac artefakty związane z transformacją układu współrzędnych
        :return:
        """
        b_num = self.btb_mul()
        b_num_extended = []
        for num, el in enumerate(b_num):  # dodaje zera zeby rozmiar sie zgadzał
            tmp = selector.extend_local_k(self.elements[num][-1], el)
            b_num_extended.append(tmp)
        return b_num_extended

    def btb_mul(self):
        """
        return multiplicztion of b array

        może zawierac artefakty związane z transformacją układu współrzędnych
        :return:
        """
        li = []
        for num, elem in enumerate(self.b):
            for nrow, row in enumerate(elem):
                for ncol, col in enumerate(row):
                    elem[nrow][ncol] = swich_poly(elem[nrow][ncol])
            self.b[num] = elem.astype('float64')
            s = self.sztywnosc(self.elements[num][-1], num)
            tmp = np.matmul(elem.T, s)

            li.append(np.matmul(tmp, elem).astype('float64'))
            # tmp = swich_poly(elem) * self.L[num]
            # if tmp is not False:
            #     li.append(np.matmul(tmp.T, tmp))
            # else:
            #     raise TypeError("łoooo, panie, całka, to bedzie płatne ekstra")
        return li

    def k_e_glob_create(self):
        glob = []
        for num, elem in enumerate(self.k_e_loc):
            # c = self.cos_oldtonew[num].T
            # v0g=np.array([0,2,1])
            # v1g = np.array(self.node_glob[num][1])
            # v2g = np.array(self.node_glob[num][2])
            # v0l = np.matmul(c.T,v0g.reshape(3,1)).reshape(3)
            # v1l = np.matmul(c.T,v1g.reshape(3,1)).reshape(3)
            # v2l = np.matmul(c.T,v2g.reshape(3,1)).reshape(3)
            # matg=np.array([v0g,v1g,v2g])
            # matl=np.array([v0l,v1l,v2l])
            # trial = macierze_num.matr_to_diag(np.array(self.elem_loc[num]), self.dof)

            c = self.cos_oldtonew[num].T
            trans = macierze_num.matr_to_diag(c, len(self.elem_loc[num]))  # rozszerza macierz do wszystkich węzłów
            transT = macierze_num.matr_to_diag(c.T, len(self.elem_loc[num]))  # rozszerza macierz do wszystkich węzłów
            # trans=np.array(reduce_matrix(trans))
            tmp = np.matmul(trans, elem)

            glob.append(np.matmul(tmp, transT))
        return glob

    def agregacja(self):
        """agreguje macierze sztywnośći elementów"""
        self.li_lrmp = []
        dim = len(self.nodes) * self.dof
        self.k = np.zeros([dim, dim])
        for n, e in enumerate(self.k_e_glob):
            temp = fu_pom.agreg_dostosowania(e, dim, *self.elements[n][:-2])  # TODO skłądnia tymczasowa
            self.li_lrmp.append(temp)
            self.k += temp

    # def k_mod_create(self):
    #     self.k_mod = np.array(self.k)
    #     som_big_val = 10 ** 10
    #     for n, e in enumerate(self.bond_cond):
    #         if e[0] == 1:
    #             if self.k_mod[n][n] == 0:
    #                 self.k_mod[n][n] = som_big_val
    #             else:
    #                 self.k_mod[n][n] *= som_big_val
    #     self.k_mod = np.array(self.k_mod)

    def k_mod_create(self):
        self.k_mod = np.copy(self.k)
        for nr, rw in enumerate(self.k_mod):
            for nc, cl in enumerate(rw):
                if self.bond_cond[nr][0] == 1 or self.bond_cond[nc][0] == 1:
                    self.k_mod[nr][nc] = 0
        for num, val in enumerate(self.bond_cond):
            if val[0] == 1:
                self.k_mod[num][num] = 1

        return 1

    def displacement(self):
        self.node_disp = np.linalg.solve(self.k_mod, self.forces)

    def reaction(self):
        self.react = np.matmul(self.k, self.node_disp) - self.forces

    def internal_forces(self):
        li=[]
        self.epsilon=[]
        disp = np.copy(self.node_disp).reshape(len(self.node_disp) // self.dof, self.dof)
        for enum, element in enumerate(self.elements):
            e_nod_disp = []
            for node in element[:-2]:
                e_nod_disp.append(disp[node])
            e_nod_disp_local = []
            for node in e_nod_disp:
                node = node.reshape((len(node), 1))  # todo sprawdzic dla pretów, bo pierdoły
                if len(node) == 3:
                    c = self.cos_oldtonew[enum]
                    e_nod_disp_local.append(
                        np.matmul(c, node)[:-1])  # robie slicing, bo przechodze na lokalny, wiec pozbywam sie z
            full_disp = np.vstack((e_nod_disp_local))

            epsilon = np.matmul(self.b[enum], full_disp)
            self.epsilon.append(epsilon)
            li.append(epsilon)
            s = self.sztywnosc(self.elements[enum][-1], enum)/self.V[enum]
            tmp =np.matmul(s, epsilon)
            tmp[1]=tmp[1]*0.8#todo *0.8
            self.local_element_stress.append(tmp)
        self.glowne()

        self.eksportuj_do_excela()
    def glowne(self):

        for elem in self.local_element_stress:
            sigx=elem[0][0]
            sigy=elem[1][0]
            tau=elem[2][0]
            dziel=(sigx+sigy)/2
            pier = math.sqrt(((sigx-sigy)/2)**2+tau**2)
            smax=dziel+pier
            smin=dziel-pier
            self.max_stress.append((smax,smin))
        return 1
    def eksportuj_do_excela(self):
        workbook = Workbook()
        sheet = workbook.active
        for rnum, row in enumerate(self.local_element_stress):
            for cnum, val in enumerate(row):
                sheet.cell(rnum + 1, cnum + 1, val[0]/1000000)
            sheet.cell(rnum + 1, 5, self.max_stress[rnum][0] / 1000000)
            sheet.cell(rnum + 1, 6, self.max_stress[rnum][1] / 1000000)


        workbook.save(filename="elementowe.xlsx")

    def show(self):
        self.print_column_matr(self.node_disp, 'przemieszczenia')
        print('\n\n')
        # self.print_column_matr(self.react, 'reakcje')

    def print_column_matr(self, matr, data_name):
        dis = np.array(matr).reshape(len(matr) // self.dof, self.dof)
        n = 0
        sum = np.array([0.0, 0.0, 0.0])

        for ndis in dis:
            print('-------------------------------------------------')
            print('{} w węźle {}'.format(data_name, n), )
            # print('x: {}'.format(round(ndis[0],4)))
            # print('y: {}'.format(round(ndis[1],4)))
            # print('z: {}'.format(round(ndis[2],4)))
            print('x: {}'.format(ndis[0]))
            print('y: {}'.format(ndis[1]))
            print('z: {}'.format(ndis[2]))
            # sum+=ndis
            if self.dof == 6:
                print('fix: {}'.format(round(ndis[3], 4)))
                print('fiy: {}'.format(round(ndis[4], 4)))
                print('fiz: {}'.format(round(ndis[5], 4)))
            n += 1
        print('suma=', sum)

        workbook = Workbook()
        sheet = workbook.active
        for rnum, row in enumerate(dis):
            for cnum, val in enumerate(row):
                if cnum < 3:  # upierdala obroty
                    sheet.cell(rnum + 1, cnum + 1, val)

        workbook.save(filename="hello_world.xlsx")

    # def inspect_input(self):
    #     """ Wyświeetla zadane dane"""
    #
    #     showlist = []
    #     for el in self.node_glob:
    #         showlist.append(('segment', (el[0], el[1]), {'color': 'g'}))
    #         showlist.append(('segment', (el[0], el[2]), {'color': 'g'}))
    #         showlist.append(('segment', (el[1], el[2]), {'color': 'g'}))
    #     if self.el_type == 'surface':
    #         tmp_forc = self.forces.reshape([int(len(self.forces) / 6), 6])[:, :3]
    #         tmp_bond = self.bond_cond.reshape([int(len(self.bond_cond) / 6), 6])[:, :3]
    #     else:
    #         tmp_forc = self.forces.reshape([int(len(self.forces) / 3), 3])
    #         tmp_bond = self.bond_cond.reshape([int(len(self.bond_cond) / 3), 3])
    #
    #     for num, nod in enumerate(self.nodes):
    #         if any(tmp_forc[num]):
    #             if nod[0] < -14:
    #                 showlist.append(('nodal_force', (nod[0:3],), {'increase': tmp_forc[num], 'scale': 0.0005, 'color':'r'}))
    #             else:
    #                 showlist.append(
    #                     ('nodal_force', (nod[0:3],), {'increase': tmp_forc[num], 'scale': 0.0005, 'color': 'b'}))
    #         if all(tmp_bond[num]):
    #             showlist.append(('kula', (nod[0:3],), {'scale': 1,'color':'black'}))
    #         elif any(tmp_bond[num]):
    #             if nod[2]!=0:
    #                 showlist.append(('kula', (nod[0:3],), {'scale': 1,'color':'chartreuse'}))
    #             else:
    #                 showlist.append(('kula', (nod[0:3],), {'scale': 1, 'color': 'fuchsia'}))
    #     tmp = visualizer.Visual(*showlist, )



    def inspect_input(self):
        """ Wyświeetla zadane dane"""
        import matplotlib.pyplot as plt
        from matplotlib.colors import ListedColormap
        import matplotlib as mpl

        def rgb_to_scale(dens=100):
            li = []
            step = 255 / dens
            for elem in range(dens):
                if elem < (dens / 2):
                    r1 = 255 - step * elem * 2
                    g1 = step * elem * 2
                    hxx = '#%02x%02x%02x' % (int(r1), int(g1), 0)
                    li.append(hxx)
                else:
                    g1 = 255 - step * (elem - dens / 2) * 2
                    b1 = step * (elem - dens / 2) * 2
                    hxx = '#%02x%02x%02x' % (0, int(g1), int(b1))
                    li.append(hxx)
            return li
        clr=rgb_to_scale()
        clr=np.flip(clr)

        #dla x,y,x
        # scale=self.node_disp.reshape(57,3)[:,0:1]
        #
        # maxdis=math.fabs(max(scale)[0])
        # mindis=math.fabs(min(scale)[0])
        # maxdis=max((maxdis,mindis))
        # mindis=-maxdis


        #dla dr
        # dx=self.node_disp.reshape(57,3)[:,0:1]
        # dy=self.node_disp.reshape(57,3)[:,1:2]
        # dz=self.node_disp.reshape(57,3)[:,2:3]
        #
        # scale=np.sqrt(dx*dx+dy*dy+dx+dz*dz)

        # scale=[]
        # for sigma in self.max_stress:
        #     sred = 1/(math.sqrt(2))*math.sqrt((sigma[0]-sigma[1])**2+sigma[0]**2+sigma[1]**2)
        #     scale.append([sred/1000000])
        #
        scale=np.array(self.max_stress)[:,1:2]/1000000



        maxdis=max(scale)[0]
        mindis= min(scale)[0]


        scope=float(maxdis-mindis)
        cmap = ListedColormap(clr)


        fig, ax = plt.subplots(figsize=(10, 8))
        fig.subplots_adjust(bottom=0.1)

        bounds = np.linspace(mindis,maxdis,20)
        norm = mpl.colors.BoundaryNorm(bounds, cmap.N, extend='both')

        fig.colorbar(mpl.cm.ScalarMappable(norm=norm, cmap=cmap)
                     , orientation='vertical',
                     label="sigma zredukowane [MPa]")
        #dla przemieszczeń
        # for nnum, node in enumerate(self.nodes):
        #     sc_val=scale[nnum]
        #     col_idx=int((sc_val-mindis)*98/scope)
        #     ax.plot(node[0], node[1], marker='o', color=clr[col_idx])
        #     ax.text(node[0] + 0.25, node[1] + 0.25, "{}".format(nnum), fontsize='small', weight=100)
        #von mises
        for enum,el in enumerate(self.node_glob):
            x=(el[0][0]+el[1][0]+el[2][0])/3
            y=(el[0][1]+el[1][1]+el[2][1])/3
            sc_val=scale[enum][0]
            col_idx=int((sc_val-mindis)*98/scope)
            ax.plot(x, y, marker='o', color=clr[col_idx])
            ax.text(x + 0.25, y + 0.25, "{}".format(enum), fontsize='small', weight=100)
        plt.show()




def rozwiaz_zadanie(dane,insp_input):
    roz = MesMain(dane)
    # return roz

if __name__ == '__main__':
    u1=rozwiaz_zadanie(prj.Kop2(), insp_input=True)
    # r1 = rozwiaz_zadanie(tst.Dane3(), insp_input=True)
    p = 0

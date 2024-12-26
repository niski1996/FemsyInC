from tkinter import Tk  # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from openpyxl import Workbook
from math import sqrt


def visual_check_excel(li):
    workbook = Workbook()
    sheet = workbook.active
    n = 0
    for group_num, group in enumerate(li):
        for num, elm in enumerate(group):
            sheet.cell(row=num + 1, column=group_num*5+1).value = elm[0]
            sheet.cell(row=num + 1, column=group_num*5+2).value = elm[1]
            sheet.cell(row=num + 1, column=group_num*5+3).value = elm[2]
            sheet.cell(row=num + 1, column=group_num*5+4).value = n
            n+=1
    workbook.save(filename='F:/zrzuty z ekranu/praca/visual_check.xlsx')


def get_path():
    """wybiera ścieżkę dostępu do pliku xls z punktami"""
    Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing
    filename = askopenfilename()  # show an "Open" dialog box and return the path to the selected file
    if valid(filename):
        return filename
    else:
        raise TypeError("FALG: MÓJ BŁĄD \n wrong extension")


def valid(str_val, ext='xlsx'):
    """check expand

    check if last three signs of given str is in list of allowed expand
    Work only for 4 letters expand (eventually with dot)

    :param
    str_val:str
    exp:array_like
        list of string
    :return
    bolean
    """
    if type(str_val) is not str:
        raise TypeError("FALG: MÓJ BŁĄD \n str_val is {}, string is required".format(type(str_val)))
    if str_val[-4:] in ext:
        return True
    else:
        return False


def exl_to_li(xls_path):
    """read xyz coords from spreadsheet and save them in list

    :param
    xsl_path:str
        path to readable excel file
    :return
    list [[x1,y1,z1],[x2,y2,z2],...]
        list of points coordinates"""
    li = []
    workbook = load_workbook(xls_path)
    sheet = workbook.active
    for n, r in enumerate(sheet):
        tmp = [float(sheet.cell(row=n + 1, column=1).value)/100,
               float(sheet.cell(row=n + 1, column=2).value)/100,
               float(sheet.cell(row=n + 1, column=3).value)/100]
        li.append(tmp)
    return li


def return_3rd(elem):
    return elem[2]


def return_3rd_of_1st(elem):
    return elem[0][2]


def return_2nd(elem):
    return elem[1]


def return_1st(elem):
    return elem[0]


def group_by_z(sorted_li):
    """ group all coord with identical z

    :param
    sorted_li : array like [[x1,y1,z1],[x2,y2,z2],...]
        list of coords sorted by z
    :return
    list
    """
    li = []
    key = None
    tmp = []
    n = 0
    sorted_li.append(['stop', 'stop', 'stop'])  # zatrzymuje iterację
    for elem in sorted_li:
        n += 1
        if key is None:
            key = elem[2]
            tmp.append(elem)
        else:
            if key == elem[2]:
                tmp.append(elem)
            else:
                li.append(tmp)
                key = elem[2]
                tmp = []
                tmp.append(elem)
    return li


def move_coords_to_first_quoater(li):
    """translate coords from li to first quater
        find lowest level of coords in gruped list, then find the farthest coords of x and y,
        take new begining of system. And translate

        :param
        li : gruped list of coord

        :return
        li : list
            list of coords translated o new system"""
    min_group = min(li, key=return_3rd_of_1st)
    min_x = min(min_group, key=return_1st)[0]
    min_y = min(min_group, key=return_2nd)[1]
    trans_vec = [-min_x, -min_y, 0]
    for gr_num, group in enumerate(li):
        for coord_num, coord in enumerate(group):
            li[gr_num][coord_num] = [sum(i) for i in list(zip(trans_vec, coord))]  # działa jak dodawanie wektorów


def sort_by_radius(group):
    """ sort elements by radius

    sort elements from one with lowest x coord, clockwise using radius form start of system.
    By default all elements should be in one z level, and arranged in circle

    :param
    group : list

    :return
    list
    """
    new_li = []
    buff = None
    target = min(group, key=return_1st)
    new_li.append(target)
    group.remove(target)
    tmp_sec=None
    for el in group:  # pierwszą iterację trzeba zrobić ręcznie, żeby być pewnym kierunku
        if el[1] < target[1]:  # sparwdza kierunkek
            second = [-el[0], -el[1], -el[2]]
            vect = [sum(i) for i in list(zip(target, second))]
            vect_len = sqrt(vect[0] ** 2 + vect[1] ** 2 + vect[2] ** 2)
            if buff is None:
                buff = vect_len
                tmp_sec = el
            else:
                if buff > vect_len:
                    buff = vect_len
                    tmp_sec = el
    buff = None
    new_li.append(tmp_sec)
    group.remove(tmp_sec)
    target = tmp_sec
    while len(group) != 0:
        for el in group:
            second = [-el[0], -el[1], -el[2]]
            vect = [sum(i) for i in list(zip(target, second))]
            vect_len = sqrt(vect[0] ** 2 + vect[1] ** 2 + vect[2] ** 2)
            if buff is None:
                buff = vect_len
                tmp_sec = el
            else:
                if buff > vect_len:
                    buff = vect_len
                    tmp_sec = el
        buff = None
        new_li.append(tmp_sec)
        group.remove(tmp_sec)
        target = tmp_sec
    return new_li

def ready_list_of_elem(e):
    """return list of elements ready to calculate"""

def ready_list_of_points(points, disp = None, forces = None):
    """return list of points ready to calculate"""
    li = []
    for num, p in enumerate(points):
        if disp is None:
            p.extend([0,0,0])
        else:
            p.extend(disp[num])
        if forces is None:
            p.extend([0,0,0])
        else:
            p.extend(forces[num])
        li.append(p)
    return li


def create_list_of_elem(li):
    """creating list of elements

    creating list of elements based on the list of elements. elements should be
    located in orbits with common center. All orbits shoulh have same amount of elements

    :param
    li : list
        list of elements grouped by level, sort by radius

    :return
    list
        list of elem coords"""
    elem_list = []
    for group_num in range(len(li)-1):
        for point_num in range(len(li[group_num])):
            elem1 = [li[group_num][point_num-1],li[group_num][point_num],li[group_num+1][point_num-1]]
            elem2 = [li[group_num][point_num],li[group_num+1][point_num-1],li[group_num+1][point_num]]
            elem_list.append(elem1)
            elem_list.append(elem2)
    return elem_list

def coord_to_index(elem, points):#TODO punkty są spakowane
    """change coords in elem to number of index"""
    points = [poi for group in points for poi in group]
    li=[]
    for e in elem:
        el_amount = len(e)
        tmp =[]
        for num in range(el_amount):
            index = points.index(e[num-1])
            tmp.append(index)
        tmp.append(0)
        tmp.append('shield')
        li.append(tmp)
    return li



lis = exl_to_li('F:/zrzuty z ekranu/praca/punkty.xlsx')
sor_li = sorted(lis, key=return_3rd)  # posortowane
group_li = group_by_z(sor_li)  # pogrupowane poziomami
move_coords_to_first_quoater(group_li)
tmp = [elem[:] for elem in group_li]
for group in group_li:
    group.sort(key=lambda x: x[1])
    group.sort(key=lambda x: x[0])

sorted_by_rad_li = list(map(sort_by_radius, group_li[:-1]))
elem_with_coords = create_list_of_elem(sorted_by_rad_li)
elem_with_index = coord_to_index(elem_with_coords, sorted_by_rad_li)
flag = 1
